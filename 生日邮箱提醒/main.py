# -*- coding: utf-8 -*-
"""
Created on Thu Dec 21 13:29:47 2017

@author: 唐小龙
"""
import random
import openpyxl
import smtplib
from email.mime.text import MIMEText
msg_from = 'xxxxxxxx@qq.com'#我的邮箱
passwd = 'xxxxxxxx'#授权码
isLogin = False
now_Email = None
usename = None

def lookInfo(Email):
    wb = openpyxl.load_workbook('Data.xlsx')
    Info = wb.get_sheet_by_name(wb.get_sheet_names()[1])
    for i in range(2,Info.max_row+1):
        if(Email == Info.cell(row = i,column = 3).value):#找到对应的邮箱了
            dic = eval(Info.cell(row = i,column = 2).value)
            for i in dic:
                s = str(dic[i])
                if('、' in s):
                    s = s.replace('、','和')
                print(s+'生日为'+i)
            wb.save('Data.xlsx')
            return 0
    print('您还没有保存任何好友的生日哟，赶紧添加吧')
def AddInfo(Email):
    dic_month = {'1': '正月', '2': '二月', '3': '三月', '4': '四月', '5': '五月', '6': '六月', '7': '七月', '8': '八月', '9': '九月', '10': '十月', '11': '十一月', '12': '十二月'}
    dic_day = {'1': '初一', '2': '初二', '3': '初三', '4': '初四', '5': '初五', '6': '初六', '7': '初七', '8': '初八', '9': '初九', '10': '初十', '11': '十一', '12': '十二', '13': '十三', '14': '十四', '15': '十五', '16': '十六', '17': '十七', '18': '十八', '19': '十九', '20': '廿十', '21': '廿一', '22': '廿二', '23': '廿三', '24': '廿四', '25': '廿五', '26': '廿六', '27': '廿七', '28': '廿八', '29': '廿九', '30': '三十', '31': '三一'}
    wb = openpyxl.load_workbook('Data.xlsx')
    Info = wb.get_sheet_by_name(wb.get_sheet_names()[1])
    flag = 0
    index = Info.max_row+1
    for i in range(2,Info.max_row+1):
        if(Email == Info.cell(row = i,column = 3).value):#找到对应的邮箱了
           index = i
           flag = 1
           break
    name = input('请输入好友的名字：')
    age = input('生日输入示例：十月初三的生日格式为：10.3：')
    month,day = age.split('.')
    if(flag == 1):
        dic = eval(Info.cell(row = index,column = 2).value)
        #{'十月初五','小明'}
        day = dic_month[month]+dic_day[day]
        if(day in dic):#已经有人是这一天生日了。
            dic[day] = dic[day]+'、'+name
        else:
            dic[day] = name
        Info.cell(row = index,column = 2).value = str(dic)
    else:
        dic = {}
        dic[dic_month[month]+dic_day[day]] = name
        Info.cell(row = index,column = 1).value = str(usename)
        Info.cell(row = index,column = 2).value = str(dic)
        Info.cell(row = index,column = 3).value = now_Email
    wb.save('Data.xlsx')
def chachong(s,n,c,info):
    wb = openpyxl.load_workbook(s)
    data = wb.get_sheet_by_name(wb.get_sheet_names()[n])
    now_index = data.max_row
    for i in range(2,now_index):
        if(info == data.cell(row = i,column = c).value):
            return 1
    return 0
def login():
    global isLogin
    global now_Email
    global usename
    while(True):
        Email = input('请输入您注册时的邮箱：')
        passwd = input('请输入密码:')
        wb = openpyxl.load_workbook('Data.xlsx')
        Register = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        flag = 0
        print(Register.max_row)
        for i in range(2,Register.max_row+1):
            if(Register.cell(row = i,column = 3).value == Email):
                if(passwd == str(Register.cell(row = i,column = 2).value)):
                    flag = 1
                    now_Email = Email
                    usename = Register.cell(row = i,column = 1).value
                    isLogin = True
                    print('登录成功')
                    break
                else:
                    print('账号或密码错误')
        if(flag == 1):
            break
    wb.save('Data.xlsx')
def register():
    global isLogin
    global now_Email
    global usename
    name = input('请输入您的用户名：')
    password = input('请输入您的密码：')
    Email = input('请输入您的邮箱：')
    while(True):
        flag = chachong('Data.xlsx',0,3,Email)
        if(flag == 0):
            break
        else:
            Email = input('此邮箱已经注册过了，请重新输入您的邮箱：')
    yanzhengma = random.randint(100000,999999)
    subject = '邮箱验证'
    content = name+'您好，您的验证码为：'+str(yanzhengma)+'请前往验证页面输入验证码完成注册！'
    msg = MIMEText(content)
    msg['Subject'] = subject
    msg['From'] = msg_from
    msg['To'] = Email
    try:
        s = smtplib.SMTP_SSL('smtp.qq.com',465)
        s.login(msg_from,passwd)
        s.sendmail(msg_from,Email,msg.as_string())
        print('发送成功，请前往邮箱查看邮件')
    except smtplib.SMTPException:
        print('邮件发送失败')
    finally:
        s.quit()
    while(True):
        s = input('请输入您接收到的验证码：')
        if(s == str(yanzhengma)):
            wb = openpyxl.load_workbook('Data.xlsx')
            Register = wb.get_sheet_by_name(wb.get_sheet_names()[0])
            now_index = Register.max_row+1
            Register.cell(row = now_index,column = 1).value = name
            Register.cell(row = now_index,column = 2).value = password
            Register.cell(row = now_index,column = 3).value = Email
            wb.save('Data.xlsx')
            isLogin = True
            now_Email = Email
            usename = name
            break
        else:
            print('验证码错误')

while(True):
    if(isLogin == False):
        select = int(input('请输入您的选择：1、登录   2、注册  3、退出\n'))
        if(select == 1):
            login()
        elif(select == 2):
            register()
        elif(select == 3):
            break
    else:
        print(usename+'您好',end = '')
        select = int((input('请输入您的选择：1、查看保存的好友生日  2、添加好友生日  3、注销当前账号\n')))
        if(select == 1):
            lookInfo(now_Email)
        elif(select == 2):
            AddInfo(now_Email)
        elif(select == 3):
            isLogin = False